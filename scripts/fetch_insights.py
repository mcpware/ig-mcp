#!/usr/bin/env python3
"""
Hourly cron job to fetch Instagram post insights.

- Fetches ALL posts from March 2026 onwards (not just latest 10)
- Stores history in SQLite (data/insights.db)
- Exports a human-readable Excel file (data/ig_insights.xlsx) every run
  with post caption, link, metrics, follower/non-follower breakdown, etc.

Usage:
    python scripts/fetch_insights.py                    # Run once
    python scripts/fetch_insights.py --since 2026-02-01 # Custom start date

Cron (every hour):
    0 * * * * cd /home/nicole/MyGithub/ig-mcp && /usr/bin/python3 scripts/fetch_insights.py >> logs/cron_insights.log 2>&1
"""

import argparse
import asyncio
import logging
import sqlite3
import sys
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, numbers
from openpyxl.utils import get_column_letter

# Add project root to path
PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

from dotenv import load_dotenv

load_dotenv(PROJECT_ROOT / ".env")

from src.config import get_settings
from src.instagram_client import InstagramClient

# ── Config ───────────────────────────────────────────────────────

DB_PATH = PROJECT_ROOT / "data" / "insights.db"
EXCEL_PATH = PROJECT_ROOT / "data" / "ig_insights.xlsx"
LOG_DIR = PROJECT_ROOT / "logs"
DEFAULT_SINCE = "2026-03-01"

LOG_DIR.mkdir(parents=True, exist_ok=True)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler(LOG_DIR / "cron_insights.log"),
    ],
)
logger = logging.getLogger(__name__)

# Metrics to fetch per post type
METRICS_IMAGE = ["impressions", "reach", "saved", "likes", "comments", "shares"]
METRICS_VIDEO = ["impressions", "reach", "saved", "likes", "comments", "shares",
                 "video_views", "plays"]


# ── Database ─────────────────────────────────────────────────────

def init_db():
    DB_PATH.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(str(DB_PATH))
    c = conn.cursor()

    c.execute("""
        CREATE TABLE IF NOT EXISTS post_insights (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            fetched_at TEXT NOT NULL,
            media_id TEXT NOT NULL,
            media_type TEXT,
            caption TEXT,
            permalink TEXT,
            posted_at TEXT,
            metric_name TEXT NOT NULL,
            metric_value INTEGER,
            UNIQUE(fetched_at, media_id, metric_name)
        )
    """)

    c.execute("""
        CREATE TABLE IF NOT EXISTS post_insights_breakdown (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            fetched_at TEXT NOT NULL,
            media_id TEXT NOT NULL,
            metric_name TEXT NOT NULL,
            breakdown_dimension TEXT NOT NULL,
            breakdown_key TEXT NOT NULL,
            breakdown_value INTEGER,
            UNIQUE(fetched_at, media_id, metric_name, breakdown_dimension, breakdown_key)
        )
    """)

    c.execute("CREATE INDEX IF NOT EXISTS idx_pi_media ON post_insights(media_id, fetched_at)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_bd_media ON post_insights_breakdown(media_id, fetched_at)")
    conn.commit()
    return conn


def store_insights(conn, fetched_at, media_id, media_type, caption, permalink,
                   posted_at, insights):
    c = conn.cursor()
    for item in insights:
        name = item.get("name", "")
        values = item.get("values", [])
        if values:
            value = values[0].get("value", 0)
            if isinstance(value, int):
                c.execute("""
                    INSERT OR REPLACE INTO post_insights
                    (fetched_at, media_id, media_type, caption, permalink, posted_at,
                     metric_name, metric_value)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """, (fetched_at, media_id, media_type, caption, permalink,
                      posted_at, name, value))
    conn.commit()


def store_breakdowns(conn, fetched_at, media_id, breakdowns):
    c = conn.cursor()
    for item in breakdowns:
        name = item.get("name", "")
        total_value = item.get("total_value", {})
        if not total_value:
            continue
        for bd in total_value.get("breakdowns", []):
            dimension = "_".join(bd.get("dimension_keys", ["unknown"]))
            for result in bd.get("results", []):
                key = "_".join(result.get("dimension_values", ["unknown"]))
                value = result.get("value", 0)
                c.execute("""
                    INSERT OR REPLACE INTO post_insights_breakdown
                    (fetched_at, media_id, metric_name, breakdown_dimension,
                     breakdown_key, breakdown_value)
                    VALUES (?, ?, ?, ?, ?, ?)
                """, (fetched_at, media_id, name, dimension, key, value))
    conn.commit()


# ── Excel Export ─────────────────────────────────────────────────

def export_excel(conn, fetched_at: str):
    """Export the latest fetch to a nicely formatted Excel file."""
    wb = Workbook()

    # ── Sheet 1: Overview (one row per post) ──
    ws = wb.active
    ws.title = "Post Overview"

    headers = [
        "Post Date", "Caption (first 80 chars)", "Type", "Link",
        "Impressions", "Reach", "Likes", "Comments", "Shares", "Saved",
        "Video Views", "Plays",
        "Reach: Follower", "Reach: Non-Follower", "Follower %",
    ]

    # Header styling
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Get all unique posts from this fetch
    c = conn.cursor()
    c.execute("""
        SELECT DISTINCT media_id, media_type, caption, permalink, posted_at
        FROM post_insights
        WHERE fetched_at = ?
        ORDER BY posted_at DESC
    """, (fetched_at,))
    posts = c.fetchall()

    for row_idx, (media_id, media_type, caption, permalink, posted_at) in enumerate(posts, 2):
        # Get metrics for this post
        c.execute("""
            SELECT metric_name, metric_value
            FROM post_insights
            WHERE fetched_at = ? AND media_id = ?
        """, (fetched_at, media_id))
        metrics = dict(c.fetchall())

        # Get breakdowns
        c.execute("""
            SELECT breakdown_key, breakdown_value
            FROM post_insights_breakdown
            WHERE fetched_at = ? AND media_id = ?
              AND breakdown_dimension = 'follow_type'
        """, (fetched_at, media_id))
        breakdowns = dict(c.fetchall())

        follower_reach = breakdowns.get("FOLLOWER", 0)
        non_follower_reach = breakdowns.get("NON_FOLLOWER", 0)
        total_breakdown = follower_reach + non_follower_reach
        follower_pct = (follower_reach / total_breakdown * 100) if total_breakdown > 0 else 0

        # Format posted_at nicely
        try:
            posted_dt = datetime.fromisoformat(posted_at.replace("Z", "+00:00"))
            posted_str = posted_dt.strftime("%Y-%m-%d %H:%M")
        except (ValueError, AttributeError):
            posted_str = posted_at or ""

        caption_short = (caption or "")[:80]

        row_data = [
            posted_str,
            caption_short,
            media_type or "",
            permalink or "",
            metrics.get("impressions", 0),
            metrics.get("reach", 0),
            metrics.get("likes", 0),
            metrics.get("comments", 0),
            metrics.get("shares", 0),
            metrics.get("saved", 0),
            metrics.get("video_views", ""),
            metrics.get("plays", ""),
            follower_reach,
            non_follower_reach,
            follower_pct,
        ]

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if col_idx == 4 and value:  # Permalink column — make it a hyperlink
                cell.hyperlink = value
                cell.font = Font(color="0563C1", underline="single")
            if col_idx == 15 and isinstance(value, (int, float)):  # Follower % column
                cell.number_format = '0.0"%"'

    # Auto-width columns
    for col_idx in range(1, len(headers) + 1):
        letter = get_column_letter(col_idx)
        max_width = max(
            len(str(ws.cell(row=r, column=col_idx).value or ""))
            for r in range(1, len(posts) + 2)
        )
        ws.column_dimensions[letter].width = min(max_width + 3, 40)

    # Freeze top row
    ws.freeze_panes = "A2"

    # ── Sheet 2: Hourly History (track changes over time) ──
    ws2 = wb.create_sheet("Hourly History")
    hist_headers = [
        "Fetched At", "Post Date", "Caption", "Media ID",
        "Impressions", "Reach", "Likes", "Comments", "Shares", "Saved",
        "Follower Reach", "Non-Follower Reach",
    ]

    for col_idx, header in enumerate(hist_headers, 1):
        cell = ws2.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Get ALL historical data (all fetches), most recent first
    c.execute("""
        SELECT pi.fetched_at, pi.posted_at, pi.caption, pi.media_id,
               MAX(CASE WHEN pi.metric_name='impressions' THEN pi.metric_value END),
               MAX(CASE WHEN pi.metric_name='reach' THEN pi.metric_value END),
               MAX(CASE WHEN pi.metric_name='likes' THEN pi.metric_value END),
               MAX(CASE WHEN pi.metric_name='comments' THEN pi.metric_value END),
               MAX(CASE WHEN pi.metric_name='shares' THEN pi.metric_value END),
               MAX(CASE WHEN pi.metric_name='saved' THEN pi.metric_value END)
        FROM post_insights pi
        GROUP BY pi.fetched_at, pi.media_id
        ORDER BY pi.fetched_at DESC, pi.posted_at DESC
    """)
    history_rows = c.fetchall()

    for row_idx, row_data in enumerate(history_rows, 2):
        fetched, posted, cap, mid, imp, rch, lik, com, shr, sav = row_data

        # Get follower breakdown for this fetch + media
        c.execute("""
            SELECT breakdown_key, breakdown_value
            FROM post_insights_breakdown
            WHERE fetched_at = ? AND media_id = ? AND breakdown_dimension = 'follow_type'
        """, (fetched, mid))
        bd = dict(c.fetchall())

        full_row = [
            fetched, posted, (cap or "")[:60], mid,
            imp or 0, rch or 0, lik or 0, com or 0, shr or 0, sav or 0,
            bd.get("FOLLOWER", 0), bd.get("NON_FOLLOWER", 0),
        ]

        for col_idx, value in enumerate(full_row, 1):
            ws2.cell(row=row_idx, column=col_idx, value=value)

    ws2.freeze_panes = "A2"
    for col_idx in range(1, len(hist_headers) + 1):
        ws2.column_dimensions[get_column_letter(col_idx)].width = 18

    # Save
    wb.save(str(EXCEL_PATH))
    logger.info(f"Excel exported to {EXCEL_PATH}")


# ── Fetch Logic ──────────────────────────────────────────────────

async def fetch_and_store(since_date: str = DEFAULT_SINCE):
    """Fetch all posts from since_date onwards, get insights, store + export."""
    settings = get_settings()
    client = InstagramClient(settings)
    conn = init_db()

    fetched_at = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
    since_dt = datetime.strptime(since_date, "%Y-%m-%d").replace(tzinfo=timezone.utc)

    try:
        await client.initialize()

        # Fetch posts in batches until we pass the since_date
        # Instagram API returns posts newest-first
        logger.info(f"Fetching posts since {since_date}...")
        all_posts = []
        batch_limit = 50  # max per request

        # Get first batch
        posts = await client.get_media_posts(limit=batch_limit)
        for post in posts:
            if post.timestamp and post.timestamp >= since_dt:
                all_posts.append(post)
            else:
                break  # Posts are newest-first, so we can stop

        logger.info(f"Found {len(all_posts)} posts since {since_date}")

        for i, post in enumerate(all_posts):
            media_id = post.id
            media_type = post.media_type.value if post.media_type else "IMAGE"
            caption = post.caption or ""
            permalink = post.permalink or ""
            posted_at = post.timestamp.isoformat() if post.timestamp else ""
            caption_preview = caption[:50].replace('\n', ' ')

            logger.info(f"  [{i+1}/{len(all_posts)}] {posted_at[:10]} | {caption_preview}...")

            # Basic metrics
            try:
                metrics = METRICS_VIDEO if media_type in ("VIDEO", "REELS") else METRICS_IMAGE
                data = await client._make_request(
                    "GET", f"{media_id}/insights",
                    params={"metric": ",".join(metrics)},
                )
                store_insights(conn, fetched_at, media_id, media_type,
                               caption[:200], permalink, posted_at,
                               data.get("data", []))
            except Exception as e:
                logger.warning(f"    Metrics failed: {e}")

            # Follower vs non-follower breakdown
            try:
                bd_data = await client._make_request(
                    "GET", f"{media_id}/insights",
                    params={
                        "metric": "reach",
                        "metric_type": "total_value",
                        "breakdown": "follow_type",
                    },
                )
                store_breakdowns(conn, fetched_at, media_id, bd_data.get("data", []))
            except Exception as e:
                logger.warning(f"    Follower breakdown failed: {e}")

            # Traffic source breakdown
            try:
                src_data = await client._make_request(
                    "GET", f"{media_id}/insights",
                    params={
                        "metric": "reach",
                        "metric_type": "total_value",
                        "breakdown": "media_product_type",
                    },
                )
                store_breakdowns(conn, fetched_at, media_id, src_data.get("data", []))
            except Exception as e:
                logger.warning(f"    Traffic source failed: {e}")

        # Export to Excel
        export_excel(conn, fetched_at)

        # Print summary
        c = conn.cursor()
        c.execute("""
            SELECT metric_name, SUM(metric_value)
            FROM post_insights WHERE fetched_at = ?
            GROUP BY metric_name ORDER BY metric_name
        """, (fetched_at,))
        logger.info("── Summary ──")
        for name, total in c.fetchall():
            logger.info(f"  {name}: {total:,}")

    except Exception as e:
        logger.error(f"Fatal error: {e}", exc_info=True)
        raise
    finally:
        conn.close()
        await client.close()


# ── Entry Point ──────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Fetch Instagram post insights")
    parser.add_argument("--since", type=str, default=DEFAULT_SINCE,
                        help=f"Fetch posts from this date onwards (default: {DEFAULT_SINCE})")
    args = parser.parse_args()

    asyncio.run(fetch_and_store(args.since))


if __name__ == "__main__":
    main()
